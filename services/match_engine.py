from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _tokens(s: Any) -> List[str]:
    s2 = _norm(s)
    if not s2:
        return []
    # tokens simples
    return [t for t in s2.split(" ") if t]


@dataclass
class MatchRow:
    desc_raw: str
    desc_norm: str
    keywords: List[str]
    herramientas: str
    materiales: str
    is_default: bool


def _load_match_rows(match_xlsx_path: str) -> Tuple[List[MatchRow], MatchRow]:
    wb = openpyxl.load_workbook(match_xlsx_path, data_only=True)
    ws = wb.active

    # Leer headers (primera fila)
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    headers_norm = [_norm(h) for h in headers]

    # Solo 3 columnas: descripción, herramientas, materiales
    def find_col(name: str) -> int:
        nn = _norm(name)
        for i, h in enumerate(headers_norm):
            if h == nn:
                return i
        raise ValueError(f"Faltan columnas en match.xlsx: '{name}'. Columnas encontradas: {headers}")

    c_desc = find_col("Descripción")
    c_tools = find_col("Herramientas")
    c_mats = find_col("Materiales")

    rows: List[MatchRow] = []
    default_row: Optional[MatchRow] = None

    for r in range(2, ws.max_row + 1):
        desc = ws.cell(row=r, column=c_desc + 1).value
        tools = ws.cell(row=r, column=c_tools + 1).value
        mats = ws.cell(row=r, column=c_mats + 1).value

        desc_str = "" if desc is None else str(desc).strip()
        desc_norm = _norm(desc_str)
        if not desc_norm:
            continue

        is_default = "default" in desc_norm  # ignora acentos por _norm()
        kw = _tokens(desc_norm)

        mr = MatchRow(
            desc_raw=desc_str,
            desc_norm=desc_norm,
            keywords=kw,
            herramientas="" if tools is None else str(tools).strip(),
            materiales="" if mats is None else str(mats).strip(),
            is_default=is_default,
        )

        if is_default and default_row is None:
            default_row = mr
        else:
            rows.append(mr)

    if default_row is None:
        raise ValueError("No se encontró fila DEFAULT en match.xlsx (Descripción debe contener 'DEFAULT').")

    if not rows:
        raise ValueError("match.xlsx no tiene filas de match (además de DEFAULT).")

    return rows, default_row


def _keyword_score(item_desc: str, pattern_keywords: List[str]) -> float:
    """
    Score = (#keywords_del_patron encontradas en el item) / (#keywords_del_patron)
    """
    item_tokens = set(_tokens(item_desc))
    patt = [k for k in pattern_keywords if k]  # limpia
    if not patt:
        return 0.0
    hit = sum(1 for k in patt if k in item_tokens)
    return hit / float(len(patt))


def enrich_items_with_match(
    items: List[Dict[str, Any]],
    match_xlsx_path: str,
    threshold: float = 0.80,
) -> List[Dict[str, Any]]:
    rows, default_row = _load_match_rows(match_xlsx_path)

    out: List[Dict[str, Any]] = []
    for it in items:
        desc = it.get("descripcion", "") or it.get("Descripción", "") or ""
        best_score = -1.0
        best_row: Optional[MatchRow] = None

        for r in rows:
            sc = _keyword_score(desc, r.keywords)
            if sc > best_score:
                best_score = sc
                best_row = r

        chosen = best_row if (best_row is not None and best_score >= threshold) else default_row

        it2 = dict(it)
        # Nombres EXACTOS que usás en el costeo
        it2["a_herramientas"] = chosen.herramientas or "herramientas de mano"
        it2["a_materiales"] = chosen.materiales or "consumibles varios"
        it2["match_score"] = float(best_score) if best_row else 0.0
        it2["match_desc"] = chosen.desc_raw
        out.append(it2)

    return out