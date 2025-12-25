from __future__ import annotations

import io
import math
import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


@dataclass
class ItemRow:
    nro: int
    descripcion: str
    unidad: str
    cantidad: float
    precio_total: int  # entero


def _norm(s: Any) -> str:
    """Normaliza: lower, sin acentos, solo alfanum + espacio."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)

    s = str(v).strip()
    if not s:
        return None

    s = s.replace("\u00a0", " ").strip()
    s = s.replace("Gs.", "").replace("Gs", "").strip()

    # manejo separadores
    if "," in s and "." in s:
        last_comma = s.rfind(",")
        last_dot = s.rfind(".")
        if last_comma > last_dot:
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) in (1, 2):
                s = parts[0].replace(".", "").replace(",", "") + "." + parts[1]
            else:
                s = s.replace(",", "")
        if "." in s:
            parts = s.split(".")
            if not (len(parts) == 2 and len(parts[1]) in (1, 2)):
                s = s.replace(".", "")

    s = re.sub(r"[^\d\.\-]", "", s)
    if not s or s in ("-", ".", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v: Any) -> Optional[int]:
    f = _to_float(v)
    if f is None:
        return None
    return int(round(f))


def _find_header_row(ws, max_scan_rows: int = 80) -> Optional[Tuple[int, Dict[str, int]]]:
    DESC_KEYS = {
        "descripcion",
        "descripciones",
        "descripcion del bien",
        "descripcion del item",
        "descripcion del ítem",
        "descripcion item",
        "desc",
    }
    TOTAL_KEYS = {
        "precio total",
        "precios totales",
        "precio total iva incluido",
        "precio total iva incl",
        "precio total (iva incluido)",
        "precio total (va incluido)",
        "precio total va incluido",
        "total",
        "totales",
        "importe total",
        "monto total",
    }
    QTY_KEYS = {"cantidad", "cant", "qty"}
    UNIT_KEYS = {"unidad", "unidad de medida", "u m", "um", "medida"}
    NRO_KEYS = {"item", "items", "nro", "no", "numero", "n", "n item"}

    def find_any(norm_headers: List[str], keys: set[str]) -> Optional[int]:
        for idx, h in enumerate(norm_headers, start=1):
            if h in keys:
                return idx
        return None

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        norm_headers = [_norm(v) for v in values]

        desc_col = find_any(norm_headers, DESC_KEYS)
        total_col = find_any(norm_headers, TOTAL_KEYS)
        if not desc_col or not total_col:
            continue

        col_map: Dict[str, int] = {
            "descripcion": desc_col,
            "precio_total": total_col,
        }
        nro_col = find_any(norm_headers, NRO_KEYS)
        qty_col = find_any(norm_headers, QTY_KEYS)
        unit_col = find_any(norm_headers, UNIT_KEYS)

        if nro_col:
            col_map["nro"] = nro_col
        if qty_col:
            col_map["cantidad"] = qty_col
        if unit_col:
            col_map["unidad"] = unit_col

        return r, col_map

    return None


def extract_items_from_excel_bytes(excel_bytes: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    meta: Dict[str, Any] = {}
    out: List[ItemRow] = []

    for ws in wb.worksheets:
        found = _find_header_row(ws)
        if not found:
            continue

        header_row, cols = found
        nro_auto = 1

        for r in range(header_row + 1, ws.max_row + 1):
            desc_val = ws.cell(row=r, column=cols["descripcion"]).value
            total_val = ws.cell(row=r, column=cols["precio_total"]).value

            desc = (str(desc_val).strip() if desc_val is not None else "").strip()
            total_int = _to_int(total_val)

            if not desc and total_int is None:
                continue
            if not desc or total_int is None:
                continue

            nro_val = ws.cell(row=r, column=cols.get("nro", 0)).value if "nro" in cols else None
            nro_int = _to_int(nro_val) if nro_val is not None else None
            if nro_int is None:
                nro_int = nro_auto
            nro_auto += 1

            unidad = ""
            if "unidad" in cols:
                u = ws.cell(row=r, column=cols["unidad"]).value
                unidad = (str(u).strip() if u is not None else "").strip()

            cantidad = 1.0
            if "cantidad" in cols:
                q = _to_float(ws.cell(row=r, column=cols["cantidad"]).value)
                if q is not None and q > 0:
                    cantidad = float(q)

            out.append(
                ItemRow(
                    nro=int(nro_int),
                    descripcion=desc,
                    unidad=unidad,
                    cantidad=cantidad,
                    precio_total=int(total_int),
                )
            )

        if out:
            break

    if not out:
        raise ValueError(
            "No se encontraron columnas clave.\n"
            "Se aceptan encabezados tipo:\n"
            "- Descripción / Descripciones / Descripción del Bien / Descripción del item\n"
            "- Precio total / Precios totales / Total / Precio total IVA incluido\n"
            "(ignorando acentos y mayúsculas)."
        )

    items = [
        {
            "nro": it.nro,
            "descripcion": it.descripcion,
            "unidad": it.unidad,
            "cantidad": it.cantidad,
            "precio_total": it.precio_total,
        }
        for it in out
    ]

    return meta, items